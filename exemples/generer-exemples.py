#!/usr/bin/env python3
"""Génère les fichiers Excel d'exemple pour l'import d'équipes.

Les fichiers étaient produits par un bloc Python collé dans un message, donc
introuvable dès qu'on voulait en ajouter un ou corriger une valeur. Un script
versionné rend la série reproductible : `python3 generer-exemples.py` réécrit
tous les fichiers à l'identique.

Deux invariants que le script garantit, et que les fichiers doivent respecter :

  * **les pseudos sont uniques d'une équipe à l'autre.** La matérialisation d'un
    import dédoublonne les joueurs par pseudo : un homonyme se retrouverait
    rattaché à deux équipes. Le script échoue plutôt que de produire un fichier
    qui créerait des données fausses ;
  * **l'effectif annoncé est respecté.** Un fichier « 4 équipes de 5 » qui en
    contiendrait 4 ne serait pas un exemple, mais un piège.

Usage :
    cd infra/exemples && python3 generer-exemples.py
"""

from __future__ import annotations

import sys

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl est requis : pip install openpyxl")

# --------------------------------------------------------------------------- #
# Effectifs
#
# Rangs plausibles du plus haut au plus bas — un roster où tout le monde est
# Radiant ne ressemble à rien et n'éprouve pas l'affichage.
# --------------------------------------------------------------------------- #

QUATRE_EQUIPES_DE_CINQ: list[tuple[str, list[tuple[str, str]]]] = [
    ("Nova Syndicate", [
        ("zephyr", "Radiant"), ("kaze", "Diamant"), ("nyx", "Diamant"),
        ("vortex", "Platine"), ("ember", "Platine"),
    ]),
    ("Iron Vanguard", [
        ("quasar", "Maître"), ("onyx", "Diamant"), ("lynx", "Diamant"),
        ("raven", "Platine"), ("pyro", "Or"),
    ]),
    ("Void Collective", [
        ("drift", "Radiant"), ("specter", "Maître"), ("havoc", "Diamant"),
        ("cipher", "Platine"), ("nomad", "Platine"),
    ]),
    ("Solaris Esports", [
        ("blitz", "Maître"), ("sable", "Diamant"), ("ronin", "Diamant"),
        ("flux", "Or"), ("talon", "Or"),
    ]),
]

DOUZE_EQUIPES_DE_TROIS: list[tuple[str, list[tuple[str, str]]]] = [
    ("Nova Syndicate", [("zephyr", "Radiant"), ("kaze", "Diamant"), ("nyx", "Diamant")]),
    ("Iron Vanguard", [("quasar", "Maître"), ("onyx", "Diamant"), ("lynx", "Platine")]),
    ("Void Collective", [("drift", "Radiant"), ("specter", "Maître"), ("havoc", "Diamant")]),
    ("Solaris Esports", [("blitz", "Maître"), ("sable", "Diamant"), ("ronin", "Or")]),
    ("Crimson Pact", [("flux", "Diamant"), ("talon", "Platine"), ("vega", "Platine")]),
    ("Northern Lights", [("orbit", "Maître"), ("shade", "Diamant"), ("pulse", "Or")]),
    ("Obsidian Order", [("brume", "Diamant"), ("silex", "Platine"), ("aster", "Platine")]),
    ("Zenith Gaming", [("cobalt", "Radiant"), ("kelvin", "Maître"), ("delta", "Diamant")]),
    ("Aurora Division", [("borealis", "Diamant"), ("frost", "Platine"), ("solstice", "Or")]),
    ("Titan Foundry", [("anvil", "Maître"), ("granit", "Diamant"), ("forge", "Platine")]),
    ("Echo Dynasty", [("reverb", "Diamant"), ("cadence", "Platine"), ("tempo", "Or")]),
    ("Phantom Ascent", [("wraith", "Radiant"), ("veil", "Maître"), ("sigil", "Diamant")]),
]


def ecrire(nom_fichier: str, entetes: list[str], lignes: list[list[str]], feuille: str) -> None:
    """Écrit un classeur d'une feuille, en-tête en gras et colonnes lisibles."""
    classeur = Workbook()
    ws = classeur.active
    ws.title = feuille
    ws.append(entetes)
    for cellule in ws[1]:
        cellule.font = Font(bold=True)
        cellule.alignment = Alignment(horizontal="left")
    for ligne in lignes:
        ws.append(ligne)
    # Un fichier qu'on ouvre pour vérifier son contenu doit être lisible sans
    # redimensionner trois colonnes à la main.
    for index, largeur in enumerate((22, 16, 14, 14, 10), start=1):
        if index <= len(entetes):
            ws.column_dimensions[get_column_letter(index)].width = largeur
    ws.freeze_panes = "A2"
    classeur.save(nom_fichier)


def lignes_par_equipe(effectifs: list[tuple[str, list[tuple[str, str]]]]) -> list[list[str]]:
    return [[equipe, pseudo, rang] for equipe, joueurs in effectifs for pseudo, rang in joueurs]


def verifier(nom_fichier: str, equipes_attendues: int, par_equipe: int, feuille: str) -> None:
    """Relit le fichier produit — on ne fait pas confiance à l'écriture."""
    ws = load_workbook(nom_fichier)[feuille]
    lignes = list(ws.values)
    noms = [ligne[0] for ligne in lignes[1:]]
    pseudos = [ligne[1] for ligne in lignes[1:]]

    if len(set(noms)) != equipes_attendues:
        sys.exit(f"{nom_fichier} : {len(set(noms))} équipes au lieu de {equipes_attendues}")
    mauvais = {nom: noms.count(nom) for nom in set(noms) if noms.count(nom) != par_equipe}
    if mauvais:
        sys.exit(f"{nom_fichier} : effectif incorrect — {mauvais}")
    if len(set(pseudos)) != len(pseudos):
        doublons = {p for p in pseudos if pseudos.count(p) > 1}
        sys.exit(f"{nom_fichier} : pseudos en double — {doublons}")

    print(
        f"  {nom_fichier} — {equipes_attendues} équipes × {par_equipe} joueurs "
        f"= {len(lignes) - 1} lignes, pseudos uniques",
    )


def main() -> None:
    print("Génération des fichiers d'exemple :")

    ecrire(
        "import-4-equipes-5-joueurs.xlsx",
        ["Équipe", "Pseudo", "Rang"],
        lignes_par_equipe(QUATRE_EQUIPES_DE_CINQ),
        feuille="Équipes",
    )
    verifier("import-4-equipes-5-joueurs.xlsx", 4, 5, "Équipes")

    ecrire(
        "import-12-equipes-3-joueurs.xlsx",
        ["Équipe", "Pseudo", "Rang"],
        lignes_par_equipe(DOUZE_EQUIPES_DE_TROIS),
        feuille="Équipes",
    )
    verifier("import-12-equipes-3-joueurs.xlsx", 12, 3, "Équipes")

    print("\nÀ importer avec le mapping Équipe A · Pseudo B · Rang C, en-tête coché.")
    print("Rappel : « participants max » compte des ÉQUIPES, pas des joueurs.")


if __name__ == "__main__":
    main()
